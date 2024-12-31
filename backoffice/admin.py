from django.contrib import admin
from django.apps import apps
from import_export.admin import ImportExportModelAdmin
from django_extensions.admin import ForeignKeyAutocompleteAdmin
from .models import Committeecalendarinused, ComComplainanttype, Contracttype, WrsSubsystemsSteps, Location,Office,CustomUser,Complaint, ComMeeting, CustomThem, Profile, Profilerole, Officestaff, Gender, Committee, Committeecalendar, Committeebranch, Committeetype, Committeesupportlocation, Zone
from .forms import CustomUserAddForm,  CustomUserForm, ProfileForm, ProfileRoleForm, OfficeForm, CommitteeForm, CommitteesupportlocationForm, ZoneForm, CommitteecalendarForm
from django.core.exceptions import ValidationError
from django.contrib import messages
from jalali_date import datetime2jalali, date2jalali
from jalali_date.admin import ModelAdminJalaliMixin, TabularInlineJalaliMixin
from admin_interface.admin import ThemeAdmin
from admin_interface.models import Theme
from django.contrib.auth.admin import UserAdmin
from django.utils.translation import gettext_lazy as _
from django.db.models.signals import post_save
from django.dispatch import receiver
from threading import local
from import_export import resources
from import_export.admin import ImportMixin
from django.core.cache import cache
from django.db import connection
from django.contrib.admin import SimpleListFilter
from django.http import HttpResponse, HttpResponseRedirect
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from django import forms
from django.shortcuts import render
from datetime import datetime
from django.forms import BaseInlineFormSet
from django.utils.safestring import mark_safe


# from django.contrib.admin import site
# import adminactions.actions as actions
#from .models import CustomThem


admin.site.site_title = "سامانه راهبری"  # Set the site title
admin.site.site_header = " سامانه راهبری"  # Set the header title
admin.site.index_title = "به سامانه راهبری خوش  آمدید"  # Set the index title



TABLES = [
    'Committeebranch', 'Committeetype'
]


def export_profiles_as_excel(modeladmin, request, queryset):
    # Create an HttpResponse object with content_type set to Excel file format
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=profiles.xlsx'

    # Create a workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Profiles"

    # Set headers for the Excel file
    headers = ['Username', 'Full Name', 'National ID', 'Mobile Number', 'Gender', 'Location']
    ws.append(headers)

    # Iterate over the queryset and add the profile data to the sheet
    for profile in queryset:
        full_name = f"{profile.firstname} {profile.lastname}"  # Combine first and last name


        # Prepare the row of data
        row = [
            profile.username,
            full_name,
            profile.NATIONALID,
            profile.mobilenumber,
            profile.gender,
            profile.location
        ]
        ws.append(row)

    # Auto adjust column widths based on the content
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook to the HttpResponse object
    wb.save(response)
    return response
export_profiles_as_excel.short_description = "خروجی با فرمت اکسل"

def export_profiles_as_pdf(modeladmin, request, queryset):
    # Prepare the HTTP response with PDF content type
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=profiles.pdf'

    # Create the PDF using ReportLab
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Set the title of the PDF
    p.setFont("Helvetica", 16)
    p.drawString(200, height - 40, "Profile Data")

    # Set the font for the content
    p.setFont("Helvetica", 12)

    # Create table headers
    p.drawString(50, height - 80, "Username")
    p.drawString(150, height - 80, "Full Name")
    p.drawString(300, height - 80, "National ID")
    p.drawString(400, height - 80, "Mobile Number")
    p.drawString(500, height - 80, "Gender")
    p.drawString(550, height - 80, "Location")

    # Iterate over the queryset and add data to the PDF
    y_position = height - 100
    for profile in queryset:
        full_name = f"{profile.firstname} {profile.lastname}"
        p.drawString(50, y_position, profile.username)
        p.drawString(150, y_position, full_name)
        p.drawString(300, y_position, str(profile.NATIONALID))
        p.drawString(400, y_position, str(profile.mobilenumber))
        p.drawString(500, y_position, str(profile.gender))
        p.drawString(550, y_position, str(profile.location))

        y_position -= 20
        if y_position < 50:
            p.showPage()  # Create a new page if content overflows
            y_position = height - 40

    p.showPage()
    p.save()
    return response
export_profiles_as_pdf.short_description = "خروجی با فرمت PDF"

class RoleFilter(SimpleListFilter):
    title = 'نقش'  # Display name in the filter sidebar
    parameter_name = 'roles'  # Query parameter in the URL

    def lookups(self, request, model_admin):
        # Provide a list of roles to filter by
        roles = Profilerole.objects.values_list('role__name', flat=True).distinct()
        return [(role, role) for role in roles]

    def queryset(self, request, queryset):
        # Apply the filter to the queryset
        if self.value():
            return queryset.filter(profileroles__role__name=self.value())
        return queryset
    

class GenderFilter(SimpleListFilter):
    title = 'جنسیت'  # Display name in the filter sidebar
    parameter_name = 'genders'  # Query parameter in the URL

    def lookups(self, request, model_admin):
        # Provide a list of roles to filter by
        names = Gender.objects.values_list('id','name')
        return [(str(g_id), g_name) for g_id, g_name in names]

    def queryset(self, request, queryset):
        # Apply the filter to the queryset
        if self.value():
            return queryset.filter(gender=self.value())
        return queryset


class ProfileResource(resources.ModelResource):
    class Meta:
        model = Profile
        # Specify the fields you want to include in the export
        fields = ('NATIONALID', 'usertype', 'firstname', 'lastname', 'location', 
                'gender', 'mobilenumber', 'isactive', 'createdby', 
                'force_password_change', 'force_profile_completion', 'password')
        batch_size = 10
        
    def get_queryset(self):
        # Use `select_related` if you are using ForeignKey fields to optimize queries
        return Profile.objects.filter(gender=1)  # Adjust as necessary
    
    def get_export_headers(self, selected_fields=None):
        # This method can return columns based on the selected fields, if provided.
        if selected_fields:
            # If specific fields are selected, only return those
            return selected_fields
        return ['NATIONALID', 'usertype', 'firstname', 'lastname', 'location', 
                'gender', 'mobilenumber', 'isactive', 'createdby', 
                'force_password_change', 'force_profile_completion', 'password']
    
    def before_export(self, queryset, **kwargs):
        return queryset.filter(gender=1)
    

# Temporary storage for the last created branch ID
_thread_locals = local()

def set_current_branch_id(branch_id):
    _thread_locals.current_branch_id = branch_id

def get_current_branch_id():
    return getattr(_thread_locals, 'current_branch_id', None)

@receiver(post_save, sender=Committeebranch)
def capture_last_created_branch_id(sender, instance, created, **kwargs):
    # Only store the ID of the branch when it's newly created
    if created:
        set_current_branch_id(instance.id)


class ProfileRoleInline(admin.TabularInline):
    model = Profilerole
    form = ProfileRoleForm
    fields = ['role']
    extra = 1
    can_delete = True


class ProfileOfficeInline(admin.TabularInline):
    model = Officestaff
    form = OfficeForm
    fields = ['office', 'role', 'committee']
    extra = 1
    can_delete = True

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        formset.form.current_user = request.user
        return formset


class CommitteebranchInline(admin.TabularInline):
    model = Committeebranch
    fields = ['code', 'name', 'isactive', 'version']
    extra = 1


class CommitteesupportlocationInline(admin.TabularInline):
    model = Committeesupportlocation
    form = CommitteesupportlocationForm
    fields = ['location', 'zone']
    extra = 3

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        formset.form.current_user = request.user
        return formset


class CommitteecalendarInline(TabularInlineJalaliMixin, admin.TabularInline):
    model = Committeecalendar
    form = CommitteecalendarForm
    fields = [
        'branch', 'availabledate', 'availabletrialcount', 
        'maxdurationofeachmeeting', 'trialcount', 'version', 'isactive', 'used_times'
    ]  # Include 'used_times' here
    readonly_fields = ['used_times']
    extra = 3    

    def get_fields(self, request, obj=None):
        if obj is None:  # If obj is None, we're adding a new instance
            return [
                'availabledate',
                'availabletrialcount',
                'maxdurationofeachmeeting',
                'trialcount',
                'version',
                'isactive'
            ]
        else:  # If obj exists, we're editing an existing instance
            return [
                'branch',
                'availabledate',
                'availabletrialcount',
                'maxdurationofeachmeeting',
                'trialcount',
                'version',
                'isactive',
                'used_times'
            ]

    def save_model(self, request, obj, form, change):
        obj.modifiedby = request.user.id
        
        #if not obj.pk:
        if not obj.createdby:
            obj.createdby = request.user.id

        super().save_model(request, obj, form, change)

    def save_new(self, form, commit=True):
            obj = super().save_new(form, commit=False)

            # Assign the captured branch ID to the 'branch' field if it's empty
            if not obj.branch:
                current_branch_id = get_current_branch_id()
                if current_branch_id:
                    obj.branch = current_branch_id  # Directly assign the branch ID
                else:
                    raise ValueError("No branch was created to assign.")

            if commit:
                obj.save()
            return obj


class BaseModelAdmin(ImportExportModelAdmin, ForeignKeyAutocompleteAdmin, admin.ModelAdmin):
    list_display = []
    search_fields = []

    def __init__(self, model, admin_site):
        list_display = [field.name for field in model._meta.fields if not field.is_relation]
        search_fields = [field.name for field in model._meta.fields if not field.is_relation]
        
        if list_display:
            self.list_display = list_display
        if search_fields:
            self.search_fields = search_fields
        
        super().__init__(model, admin_site)

models = apps.get_containing_app_config("backoffice").get_models()

for model in models:
    if model.__name__ != 'Profile' and model.__name__ != 'Profilerole' and model.__name__ != 'Office' and model.__name__ != 'Committeebranch' and model.__name__ != 'Committeetype'and model.__name__ in TABLES:
        try:
            admin.site.register(model, type(f'{model.__name__}Admin', (BaseModelAdmin,), {}))
        except admin.sites.AlreadyRegistered:
            pass
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error registering {model.__name__}: {e}")


# Custom access level filtering for each ModelAdmin
class CustomAdminAccessMixin:
    def has_module_permission(self, request):
        # Retrieve the access level from the user's profile
        if isinstance(request.user, Profile):
            access_level = getattr(request.user, 'admin_access_level', 0)

            # Check visibility based on the model and access level
            if self.model == Profile and access_level in [1, 2, 3]:
                return True
            elif self.model == Committee and access_level == 3:
                return True
            elif self.model == CustomThem and access_level == 1:
                return True
            elif self.model == CustomUser and access_level in [2, 3]:
                return True
            elif access_level == 0:
                return True
            return False
        elif request.user.is_superuser:
            return True
        return False

    def get_model_perms(self, request):
        """Only return model permissions if user has access."""
        if self.has_module_permission(request):
            return super().get_model_perms(request)
        return {}
    

class ProfileAdmin(ImportMixin, CustomAdminAccessMixin, admin.ModelAdmin):
    form = ProfileForm
    actions = [export_profiles_as_pdf, export_profiles_as_excel]
    resource_class = ProfileResource

    list_display = [
        'firstname', 'lastname', 'NATIONALID', 'get_gender', 'get_loc', 'get_roles', 
        'mobilenumber', 'isactive', 'createdby', 
        'force_password_change', 'force_profile_completion', 'DJANGO_IS_ACTIVE', 'DJANGO_IS_STAFF', 'DJANGO_IS_SUPERUSER', 'DJANGO_LAST_LOGIN'
    ]

    search_fields = [
        'firstname__icontains', 
        'lastname__icontains', 
        'NATIONALID__exact', 
        'mobilenumber__exact', 
        'username__exact'
    ]

    readonly_fields = ['createdat', 'createdby', 'modifiedat', 'modifiedby']
    
    list_filter = (RoleFilter, GenderFilter)
    list_per_page = 20
    #show_full_result_count = False
    
    inlines = [ProfileRoleInline, ProfileOfficeInline]

    class Media:
        js = ('admin_profile.js', 'nationalid_check.js',)

    def get_roles(self, obj):
        # Get all related roles, return '-' if there are none
        roles = [str(role.role) for role in obj.profileroles.all()]  # Assuming role has a `__str__` method
        return ", ".join(roles) if roles else '-'

    get_roles.short_description = 'نقش'

    def get_loc(self, obj):
        cache_key = f"location_{obj.location}"
        location_name = cache.get(cache_key)
        if not location_name:
            with connection.cursor() as cursor:
                cursor.execute("SELECT name FROM location WHERE id = %s", [obj.location])
                result = cursor.fetchone()
                location_name = result[0] if result else ''
                cache.set(cache_key, location_name, timeout=3600)  # Cache for 1 hour
        return location_name
    get_loc.short_description = 'نام شهر'

    def get_gender(self, obj):
        try:
            gen_name = Gender.objects.get(id=obj.gender).name            
        except Gender.DoesNotExist:
            gen_name = ''
        return gen_name
    get_gender.short_description = 'جنسیت'
    
    def get_form(self, request, obj=None, **kwargs):
        # Pass the current user to the form
        kwargs['form'] = ProfileForm
        form = super().get_form(request, obj, **kwargs)
        form.current_user = request.user  # Set the user
        return form
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)

        # If the user is a superuser, check the province or profile location
        if request.user.is_superuser:
            # Check if `CustomUser` has a province
            if hasattr(request.user, 'province') and request.user.province is not None:
                return qs.filter(location=request.user.province)  # Filter based on user's province

            # Access the Profile model manually, using the user (assuming Profile has a `user` FK or related logic)
            try:
                profile = Profile.objects.get(username=request.user.username)  # Assuming the `Profile` model has a `username` field matching `CustomUser.username`
                #profile = Profile.objects.filter(username=request.user.username).first()
                if profile.location is not None:
                    return qs.filter(location=profile.location, usertype__code="03")  # Filter based on profile location
            except Profile.DoesNotExist:
                pass  # Profile doesn't exist, so continue as no filtering is done

            return qs  # If neither province nor location is set, show all

        # For other user logic, you can add custom conditions for groups, etc.
        return qs.none()

    def get_fields(self, request, obj=None):
        if obj:  # If obj is not None, this is an update view
            return [
                'NATIONALID', 'usertype', 'firstname', 'lastname', 'location', 
                'gender', 'mobilenumber', 'isactive', 'createdby', 
                'force_password_change', 'force_profile_completion', 'password',
                'deathdate', 'activefrom', 'activeto', 'childrennumber',
                'createdat', 'defunctdate', 'foundationdate', 'graduation', 
                'maritalstatus', 'modifiedat', 'modifiedby', 'religion', 
                'postalcode', 'zonecode', 'FOREIGNERID', 'faxnumber', 
                'nationalcertificatecode', 'nationalcertificateserial', 
                'phonenumber', 'fathername', 'organizationname', 'username', 
                'address', 'secretkey', 'birthdate', 'currenttaminworkshopcode', 
                'email', 'graduationlevel', 'physicalcondition', 'idno', 
                'ispermittedtoreceivebyemail', 'ispermittedtoreceivebyfax', 
                'militaryservicestatus', 'parentlocation', 'nationality', 
                'officialemail', 'externalusertype', 'is_converted', 
                'birth_place', 'DJANGO_IS_ACTIVE', 'DJANGO_IS_STAFF', 'DJANGO_IS_SUPERUSER'
            ]
        else:

            return [
                'NATIONALID', 'usertype', 'firstname', 'lastname', 'username', 'location', 
                'gender', 'mobilenumber', 'isactive', 'createdby', 
                'force_password_change', 'force_profile_completion', 'password'
            ]        
       
    def save_model(self, request, obj, form, change):
        """
        Override the save_model method to catch validation errors from the model
        and prevent success messages when there are errors.
        """
        try:
            if not obj.pk:  # This means it's a new object being created
                if not obj.createdby:
                    obj.createdby = request.user.id  # Set createdby to the current user

            obj.modifiedby = request.user.id

            # Attempt to save the object
            super().save_model(request, obj, form, change)

        except ValidationError as e:
            # Catch validation errors and show them on the form
            form.add_error(None, e)  # Adds the error to the form's non-field errors
            messages.set_level(request, messages.ERROR)
            messages.add_message(request, messages.ERROR, messages.error(request, e.message))  # Optionally show a Django message
            return

    def save_formset(self, request, form, formset, change):
        """
        Override save_formset to handle deletions explicitly.
        """
        if formset.model == Profilerole or formset.model == Officestaff:
            # Check for any deletions
            instances = formset.save(commit=False)
            for obj in formset.deleted_objects:
                obj.delete()  # Handle deletion
            for instance in instances:
                instance.save()
        else:
            super().save_formset(request, form, formset, change)
                     

class ProfileRoleAdmin(admin.ModelAdmin):
    form = ProfileRoleForm
    list_display = [
        'profile', 'role', 'isactive', 'createdat', 'createdby', 'modifiedat',
        'modifiedby', 'is_default'
    ]
    search_fields = [
        'profile__firstname', 'profile__lastname', 'role__name'
    ]
    readonly_fields = ['createdat', 'createdby', 'modifiedat', 'modifiedby', 'isactive']

    def get_fields(self, request, obj=None):
        if obj:  # If obj is not None, this is an update view
            return [
               'role',
            ]
        else:

            return [
         'profile', 'role', 'isactive', 'createdat', 'createdby', 'modifiedat',
        'modifiedby', 'is_default'
            ]
    

class OfficeAdmin(admin.ModelAdmin):
    form = OfficeForm
    list_display = [
        'code', 'address', 'isactive', 'ishqoffice', 'name',
        'parentid', 'economical_code', 'english_name', 'fax_number',
        'location_id', 'national_code', 'registry_date', 'zone_id',
        'phone_number', 'office_role', 'ownership', 'municipal_zone_code'
    ]
    search_fields = [
        'code', 'address', 'name', 'economical_code', 'national_code'
    ]
    readonly_fields = ['createdat', 'createdby', 'modifiedat', 'modifiedby']

    def get_form(self, request, obj=None, **kwargs):
        
        kwargs['form'] = OfficeForm
        form = super().get_form(request, obj, **kwargs)
        form.current_user = request.user  # Set the user
        return form

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:  # Superuser can see all offices
            return qs
        elif request.user.groups.filter(name__startswith='CanCreateUser-').exists():
            allowed_locations = [group.name.split('-')[1] for group in request.user.groups.all() if group.name.startswith('CanCreateUser-')]
            return qs.filter(location__name__in=allowed_locations)
        return qs.none() 


class CommitteeAdmin(ImportExportModelAdmin, ModelAdminJalaliMixin, CustomAdminAccessMixin, admin.ModelAdmin):
    form = CommitteeForm
    list_display = ['code', 'name', 'commiteetype', 'office', 'address', 'createdat', 'createdby', 'isactive',
                    'modifiedat', 'modifiedby',  'version']
    readonly_fields = ['createdat', 'createdby', 'modifiedat', 'modifiedby']
    list_display_links = ["code","name","commiteetype", "office"]
    list_filter = ('name', 'commiteetype')
    list_per_page = 30

    inlines = [CommitteebranchInline, CommitteecalendarInline, CommitteesupportlocationInline]

    def get_form(self, request, obj=None, **kwargs):
        
        kwargs['form'] = CommitteeForm
        form = super().get_form(request, obj, **kwargs)
        form.current_user = request.user  # Set the user
        return form
   
    def save_model(self, request, obj, form, change):
        if form.cleaned_data.get('office'):
            obj.office = form.cleaned_data['office']
        obj.modifiedby = request.user.id
        if not obj.pk:
            if not obj.createdby:
                    obj.createdby = request.user.id
            
        super().save_model(request, obj, form, change)

    def get_queryset(self, request):
        """
        Modify the queryset to filter committees based on the user's location or province.
        """
        queryset = super().get_queryset(request)
        
        # If the user has a province or location, filter the committees
        if request.user.is_superuser:
            if hasattr(request.user, 'province') and request.user.province is not None:
                return queryset.filter(office__location_id=request.user.province)
            try:
                profile = Profile.objects.get(username=request.user.username)
                user_offices = Officestaff.objects.filter(profile=profile, isactive=True).values_list('office', flat=True)
                if user_offices:
                    return queryset.filter(office_id__in=user_offices)
                else:
                    return queryset.none()
                # if profile.location is not None:
                #     return queryset.filter(office__location_id=profile.location)
            except Profile.DoesNotExist:
                return queryset
      
            return queryset
        return queryset.none()


class CommitteebranchAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'committee', 'get_comms' ,'createdat', 'createdby', 'isactive', 'modifiedat', 'modifiedby', 'version']

    def get_comms(self, obj):
        # Get all related roles, return '-' if there are none
        comms = [str(comm.name) for comm in obj.Committee.all()]  # Assuming role has a `__str__` method
        return ", ".join(comms) if comms else '-'

    get_comms.short_description = 'کمیته'

    def save_model(self, request, obj, form, change):
        obj.modifiedby = request.user.id

        if not obj.pk:
            if not obj.createdby:
                    obj.createdby = request.user.id
        super().save_model(request, obj, form, change)


class CommitteecalendarAdmin(ModelAdminJalaliMixin, admin.ModelAdmin):

    form = CommitteecalendarForm
    list_display = ['committee', 'branch', 'get_availabledate_jalali', 'maxdurationofeachmeeting', 'createdat', 'createdby', 
                    'modifiedat', 'modifiedby', 'trialcount', 'version', 'isactive', 'availabletrialcount']
    
    fields = [
        'committee', 'branch', 'availabledate', 'availabletrialcount', 
        'maxdurationofeachmeeting', 'trialcount', 'version', 'isactive', 'used_times'
    ]

    readonly_fields = ['used_times']

    def get_fieldsets(self, request, obj=None):
        fieldsets = super().get_fieldsets(request, obj)
        if not obj:  # Only modify when creating a new instance
            fieldsets.insert(0, (
                "Committee Information",  # Header title
                {
                    'fields': ('branch', 'availabledate', 'availabletrialcount', 'maxdurationofeachmeeting', 'trialcount', 'version', 'isactive'),
                    'description': 'This section contains the committee and scheduling details.'
                }
            ))
        return fieldsets

  
    @admin.display(description='تاریخ و زمان دردسترس', ordering='createdat')
    def get_availabledate_jalali(self, obj):
        return datetime2jalali(obj.availabledate).strftime('%a, %d %b %Y %H:%M:%S')    

    def save_model(self, request, obj, form, change):
        obj.modifiedby = request.user.id        
        #if not obj.pk:
        if not obj.createdby:
            obj.createdby = request.user.id
        super().save_model(request, obj, form, change)

    def get_form(self, request, obj=None, **kwargs):
        """
        Modify the form based on whether the fields should be read-only or editable.
        """
        form = super().get_form(request, obj, **kwargs)

        if obj:
            # Check if the available date is in the future and if the Committeecalendar.id is not in Committeecalendarinused
            if obj.availabledate <= datetime.now() or Committeecalendarinused.objects.filter(committeecalendar=obj.pk).exists():
                # Make fields read-only
                for field in form.base_fields:
                    form.base_fields[field].widget.attrs['readonly'] = True
        return form


class CommitteesupportlocationAdmin(admin.ModelAdmin):

    form = CommitteesupportlocationForm

    def get_form(self, request, obj=None, **kwargs):
        
        kwargs['form'] = CommitteesupportlocationForm
        form = super().get_form(request, obj, **kwargs)
        form.current_user = request.user  # Set the user
        return form

    def save_model(self, request, obj, form, change):
        obj.modifiedby = request.user.id
        
        if not obj.pk:
            if not obj.createdby:
                    obj.createdby = request.user.id
        super().save_model(request, obj, form, change)


class ZoneAdmin(admin.ModelAdmin):

    list_display = [
        'code', 'name', 'location', 'createdat',  'createdby',  'modifiedat', 'modifiedby'
        ]

    form=ZoneForm

    def get_form(self, request, obj=None, **kwargs):
        
        kwargs['form'] = ZoneForm
        form = super().get_form(request, obj, **kwargs)
        form.current_user = request.user  # Set the user
        return form


class CustomUserAdmin(CustomAdminAccessMixin, UserAdmin):
    # Define the fields to display in the user list view
    add_form = CustomUserAddForm
    form = CustomUserForm
    list_display = ('username', 'email', 'first_name', 'last_name', 'province', 'admin_access_level', 'is_staff')
    

    # Add fieldsets for the user detail view
    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        (_('Personal info'), {'fields': ('first_name', 'last_name', 'email', 'province', 'admin_access_level')}),
        (_('Permissions'), {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        (_('Important dates'), {'fields': ('last_login', 'date_joined')}),
    )

    # Specify fields to include in the add user form
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('username', 'password1', 'password2', 'email', 'province'),
        }),
    )

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.user = request.user  # Pass the request user
        return form
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
             if not hasattr(request.user, 'province') or request.user.province is None:  # Superuser can see all profiles
                return qs
             return qs.filter(province=request.user.province)
        elif request.user.groups.filter(name__startswith='CanCreateUser-').exists():
            allowed_locations = [group.name.split('-')[1] for group in request.user.groups.all() if group.name.startswith('CanCreateUser-')]
            return qs.filter(location__name__in=allowed_locations)
        return qs.none()

  
class ComplaintAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    list_display = ['get_complainanttype', 'jobtitle', 'get_contracttype', 'createdat', 'filledbycomplainant',
    'finalsalary', 'firstviewedat', 'firstviewedby', 'hasinsurance', 'modifiedat', 'canceled',
    'modifiedby' , 'paymentmethod', 'get_profile', 'requestcode', 'requestdescription', 'requesttitle',
    'salary', 'state', 'step', 'get_step', 'tamininsurancenumber', 'taminworkshopcode' , 'totalexperiencesday',
    'totalexperiencesmonth', 'version', 'workshopaddress', 'workshopcertificateid', 'workshopemail',
    'workshopexperiencesday', 'workshopexperiencesmonth', 'workshopfax', 'workshopguildnumber', 'workshoplegalid',
    'get_location', 'workshopmobile', 'workshopname', 'workshopphone' , 'workshoppostalcode', 'workshopzone',
]
    list_per_page = 10
    list_filter = ('state', 'step', 'createdat')

    def get_step(self, obj):
        try:
            step_name = WrsSubsystemsSteps.objects.get(step_full_code=obj.step + obj.state)            
        except Gender.DoesNotExist:
            step_name = ''
        return step_name
    get_step.short_description = 'گام'

    def get_contracttype(self, obj):
        try:
            contracttype_name = Contracttype.objects.get(id=obj.contracttype).name         
        except Contracttype.DoesNotExist:
            contracttype_name = ''
        return contracttype_name
    get_contracttype.short_description = 'نوع قرارداد'

    def get_complainanttype(self, obj):
        try:
            contracttype_name = ComComplainanttype.objects.get(id=obj.complainanttype).name         
        except ComComplainanttype.DoesNotExist:
            contracttype_name = ''
        return contracttype_name
    get_complainanttype.short_description = 'نوع خواهان'

    def get_profile(self, obj):
        try:
            profile_name = Profile.objects.get(id=obj.profile).firstname  + ' ' + Profile.objects.get(id=obj.profile).lastname
        except Profile.DoesNotExist:
            profile_name = ''
        return profile_name
    get_profile.short_description = 'نام و نام خانوادگی خواهان'

    def get_location(self, obj):
        try:
            location_name = Location.objects.get(id=obj.workshoplocation)
        except Location.DoesNotExist:
            location_name = ''
        return location_name
    get_location.short_description = 'شهر/استان کارگاه'



# Register the Profile model with the ProfileAdmin
admin.site.register(Profile, ProfileAdmin)
admin.site.register(Committee, CommitteeAdmin)
#admin.site.register(Committeecalendar, CommitteecalendarAdmin)
#admin.site.register(Committeesupportlocation, CommitteesupportlocationAdmin)
#admin.site.register(Zone, ZoneAdmin)
admin.site.unregister(Theme)
admin.site.register(CustomThem, ThemeAdmin)
#admin.site.register(ComMeeting)
#admin.site.register(Complaint, ComplaintAdmin)
admin.site.register(CustomUser, CustomUserAdmin)
#admin.site.register(Officestaff)
#admin.site.register(Office, OfficeAdmin)
#actions.add_to_site(site)